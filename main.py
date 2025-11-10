# =====================================================
#  IP Intelligence Aggregator - Backend API
#  Autor: Felipe Prates
#  Versão: 3.4 (ajustes de timeout /api/insights)
#  Descrição: Backend FastAPI para consultas de IP,
#  Domínio e Hash, CSV/JSON e controle via config.py.
#  Última atualização: 24/10/2025
# =====================================================

from pathlib import Path
from typing import List, Tuple
from fastapi import Body, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openai import OpenAI
from config import settings
import asyncio
import csv
import io
import json
import re
import socket
import httpx

app = FastAPI(title="ASPER IP Intelligence", version="3.4")

# ====== Static / Index ======
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
INDEX_CANDIDATES = [STATIC_DIR / "index.html", BASE_DIR / "index.html"]

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.get("/", include_in_schema=False)
async def index():
    for p in INDEX_CANDIDATES:
        if p.exists():
            return FileResponse(str(p))
    return Response("index.html não encontrado.", status_code=404)

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    fav = STATIC_DIR / "favicon.png"
    if fav.exists():
        return FileResponse(str(fav))
    return Response(status_code=204)

@app.get("/health", include_in_schema=False)
async def health():
    return {"status": "ok", "env": settings.ENV}

# ====== HTTP client ======
def _client() -> httpx.AsyncClient:
    timeout = httpx.Timeout(settings.REQUEST_TIMEOUT_SECONDS,
                            connect=settings.CONNECT_TIMEOUT_SECONDS)
    return httpx.AsyncClient(timeout=timeout, follow_redirects=True)

# ===================== LOOKUPS (IP) ===================== #
async def vt_lookup(client: httpx.AsyncClient, ip: str):
    url = f"{settings.VT_BASE}/ip_addresses/{ip}"
    r = await client.get(url, headers={"x-apikey": settings.VT_API_KEY})
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    attrs = data.get("data", {}).get("attributes", {})
    stats = attrs.get("last_analysis_stats", {})
    return {
        "found": True,
        "malicious": stats.get("malicious"),
        "suspicious": stats.get("suspicious"),
        "harmless": stats.get("harmless"),
        "undetected": stats.get("undetected"),
        "last_analysis_date": attrs.get("last_analysis_date"),
        "country": attrs.get("country"),
        "asn": attrs.get("as_owner"),
        "raw": data,
    }

async def abuseipdb_lookup(client: httpx.AsyncClient, ip: str):
    url = f"{settings.ABUSEIPDB_BASE}/check"
    params = {"ipAddress": ip, "maxAgeInDays": settings.MAX_AGE_DAYS_ABUSEIPDB, "verbose": True}
    r = await client.get(url, headers={"Key": settings.ABUSEIPDB_API_KEY, "Accept": "application/json"}, params=params)
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json().get("data", {})
    return {
        "found": True,
        "abuseConfidenceScore": data.get("abuseConfidenceScore"),
        "totalReports": data.get("totalReports"),
        "lastReportedAt": data.get("lastReportedAt"),
        "countryCode": data.get("countryCode"),
        "isp": data.get("isp"),
        "usageType": data.get("usageType"),
        "domain": data.get("domain"),
        "raw": {"data": data},
    }

async def ipqs_lookup(client: httpx.AsyncClient, ip: str):
    url = f"https://ipqualityscore.com/api/json/ip/{settings.IPQS_API_KEY}/{ip}"
    r = await client.get(url)
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    return {
        "found": True,
        "fraud_score": data.get("fraud_score"),
        "proxy": data.get("proxy"),
        "vpn": data.get("vpn"),
        "tor": data.get("tor"),
        "recent_abuse": data.get("recent_abuse"),
        "ISP": data.get("ISP"),
        "organization": data.get("organization"),
        "country_code": data.get("country_code"),
        "city": data.get("city"),
        "region": data.get("region"),
        "ASN": data.get("ASN"),
        "raw": data,
    }

async def ipinfo_lookup(client: httpx.AsyncClient, ip: str):
    url = f"https://ipinfo.io/{ip}?token={settings.IPINFO_API_KEY}"
    r = await client.get(url)
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    return {
        "found": True,
        "ip": data.get("ip"),
        "org": data.get("org"),
        "hostname": data.get("hostname"),
        "country": data.get("country"),
        "region": data.get("region"),
        "city": data.get("city"),
        "loc": data.get("loc"),
        "anycast": data.get("anycast"),
        "raw": data,
    }

async def otx_lookup(client: httpx.AsyncClient, ip: str):
    url = f"{settings.OTX_BASE}/indicators/IPv4/{ip}/general"
    r = await client.get(url, headers={"X-OTX-API-KEY": settings.OTX_API_KEY})
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    pulse_count = len(data.get("pulse_info", {}).get("pulses", []) or [])
    return {
        "found": True,
        "alexa": data.get("alexa"),
        "asn": data.get("asn"),
        "country_code": data.get("country_code"),
        "pulse_count": pulse_count,
        "raw": data,
    }

# ===================== LOOKUPS (DOMÍNIO) ===================== #
def resolve_domain_sync(domain: str) -> list[str]:
    try:
        _, _, ips = socket.gethostbyname_ex(domain)
        return ips or []
    except Exception:
        return []

async def resolve_domain(domain: str) -> list[str]:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, resolve_domain_sync, domain)

async def vt_domain_lookup(client: httpx.AsyncClient, domain: str):
    url = f"{settings.VT_BASE}/domains/{domain}"
    r = await client.get(url, headers={"x-apikey": settings.VT_API_KEY})
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    attrs = data.get("data", {}).get("attributes", {})
    stats = attrs.get("last_analysis_stats", {})
    cats = (attrs.get("categories") or {})
    return {
        "found": True,
        "reputation": attrs.get("reputation"),
        "categories": list(cats.values()),
        "harmless": stats.get("harmless"),
        "malicious": stats.get("malicious"),
        "suspicious": stats.get("suspicious"),
        "undetected": stats.get("undetected"),
        "raw": data,
    }

async def otx_domain_lookup(client: httpx.AsyncClient, domain: str):
    url = f"{settings.OTX_BASE}/indicators/domain/{domain}/general"
    r = await client.get(url, headers={"X-OTX-API-KEY": settings.OTX_API_KEY})
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    pulse_count = len(data.get("pulse_info", {}).get("pulses", []) or [])
    whois = data.get("whois", {})
    return {
        "found": True,
        "alexa": data.get("alexa"),
        "pulse_count": pulse_count,
        "whois": {
            "registrar": whois.get("registrar"),
            "creation_date": whois.get("creation_date"),
            "expiration_date": whois.get("expiration_date"),
        },
        "raw": data,
    }

# ===================== LOOKUPS (HASH) ===================== #
HEX_RE = re.compile(r"^[A-Fa-f0-9]+$")

def _hash_kind(h: str) -> str:
    h = h.strip().lower()
    if HEX_RE.match(h) and len(h) == 32:
        return "md5"
    if HEX_RE.match(h) and len(h) == 64:
        return "sha256"
    return ""

async def vt_file_lookup(client: httpx.AsyncClient, h: str):
    url = f"{settings.VT_BASE}/files/{h}"
    r = await client.get(url, headers={"x-apikey": settings.VT_API_KEY})
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    attrs = data.get("data", {}).get("attributes", {})
    stats = attrs.get("last_analysis_stats", {})
    return {
        "found": True,
        "type_description": attrs.get("type_description"),
        "size": attrs.get("size"),
        "first_submission_date": attrs.get("first_submission_date"),
        "last_analysis_date": attrs.get("last_analysis_date"),
        "md5": attrs.get("md5"),
        "sha256": attrs.get("sha256"),
        "names": attrs.get("names"),
        "reputation": attrs.get("reputation"),
        "stats": stats,
        "raw": data,
    }

async def ha_lookup(client: httpx.AsyncClient, h: str):
    if not getattr(settings, "HYBRID_API_KEY", None):
        return {"error": "missing_api_key"}
    url = f"{settings.HYBRID_BASE}/overview/{h}"
    headers = {
        "api-key": settings.HYBRID_API_KEY,
        "user-agent": "Falcon Sandbox",
        "accept": "application/json",
    }
    r = await client.get(url, headers=headers)
    if r.status_code == 404:
        return {"found": False}
    r.raise_for_status()
    data = r.json()
    item = (data[0] if isinstance(data, list) and data else data) or {}
    return {
        "found": True,
        "verdict": item.get("verdict"),
        "threat_score": item.get("threat_score"),
        "vx_family": item.get("vx_family"),
        "type": item.get("type"),
        "size": item.get("size"),
        "md5": item.get("md5"),
        "sha256": item.get("sha256"),
        "analysis_start_time": item.get("analysis_start_time"),
        "submitted": item.get("submission_type"),
        "raw": data,
    }

# ===================== ENDPOINTS ===================== #
@app.get("/api/lookup")
async def lookup_ip(ip: str = Query(..., description="Endereço IPv4/IPv6")):
    if len(ip) > 100 or " " in ip:
        raise HTTPException(status_code=400, detail="IP inválido.")
    async with _client() as c:
        tasks = [vt_lookup(c, ip), abuseipdb_lookup(c, ip), ipqs_lookup(c, ip), ipinfo_lookup(c, ip), otx_lookup(c, ip)]
        res = await _gather_safe(*tasks)
    return JSONResponse({"ip": ip, "sources": {
        "virustotal": res[0], "abuseipdb": res[1], "ipqualityscore": res[2], "ipinfo": res[3], "otx": res[4]
    }})

@app.get("/api/lookup/domain")
async def lookup_domain(domain: str = Query(..., description="Domínio (ex.: exemplo.com)")):
    if len(domain) > 255 or " " in domain:
        raise HTTPException(status_code=400, detail="Domínio inválido.")
    async with _client() as c:
        ips = await resolve_domain(domain)
        tasks = [vt_domain_lookup(c, domain), otx_domain_lookup(c, domain)]
        dom_res = await _gather_safe(*tasks)
    return JSONResponse({"domain": domain, "resolved_ips": ips, "sources": {
        "virustotal": dom_res[0], "otx": dom_res[1]
    }})

@app.post("/api/lookup/batch")
async def lookup_batch(payload: dict = Body(..., example={"ips": ["1.1.1.1","8.8.8.8"], "domains": ["exemplo.com"]})):
    ips = payload.get("ips") or []
    domains = payload.get("domains") or []
    if not isinstance(ips, list) or not isinstance(domains, list):
        raise HTTPException(status_code=400, detail="Formato inválido. Use listas para 'ips' e 'domains'.")

    sem = asyncio.Semaphore(5)

    async def do_ip(ip: str):
        async with sem:
            async with _client() as c:
                res = await _gather_safe(vt_lookup(c, ip), abuseipdb_lookup(c, ip),
                                         ipqs_lookup(c, ip), ipinfo_lookup(c, ip), otx_lookup(c, ip))
                return {"ip": ip, "sources": {
                    "virustotal": res[0], "abuseipdb": res[1], "ipqualityscore": res[2], "ipinfo": res[3], "otx": res[4]
                }}

    async def do_domain(d: str):
        async with sem:
            async with _client() as c:
                ips_res = await resolve_domain(d)
                res = await _gather_safe(vt_domain_lookup(c, d), otx_domain_lookup(c, d))
                return {"domain": d, "resolved_ips": ips_res, "sources": {"virustotal": res[0], "otx": res[1]}}

    ip_tasks = [do_ip(i.strip()) for i in ips if i and isinstance(i, str)]
    dom_tasks = [do_domain(d.strip()) for d in domains if d and isinstance(d, str)]

    ip_results, dom_results = [], []
    if ip_tasks:
        ip_results = await asyncio.gather(*ip_tasks, return_exceptions=False)
    if dom_tasks:
        dom_results = await asyncio.gather(*dom_tasks, return_exceptions=False)

    return JSONResponse({"count": {"ips": len(ip_results), "domains": len(dom_results)},
                         "ips": ip_results, "domains": dom_results})

# ===== /api/lookup/hash =====
@app.get("/api/lookup/hash")
async def lookup_hash(hash: str = Query(..., description="MD5 (32 hex) ou SHA256 (64 hex)")):
    kind = _hash_kind(hash)
    if not kind:
        raise HTTPException(status_code=400, detail="Hash inválido. Use MD5 (32 hex) ou SHA256 (64 hex).")
    async with _client() as c:
        vt_res, ha_res = await _gather_safe(vt_file_lookup(c, hash), ha_lookup(c, hash))
    return JSONResponse({"hash": hash, "kind": kind, "sources": {
        "virustotal_file": vt_res, "hybrid_analysis": ha_res
    }})

# ===================== Upload CSV/XLSX ===================== #
IP_RE = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$|^[0-9a-fA-F:]+$")
DOMAIN_RE = re.compile(r"^(?=.{1,255}$)([a-zA-Z0-9-_]+\.)+[a-zA-Z]{2,}$")

def _auto_detect_cells(row: List[str], ips: List[str], doms: List[str]):
    for cell in row:
        v = str(cell).strip()
        if not v:
            continue
        if IP_RE.match(v):
            ips.append(v)
        elif DOMAIN_RE.match(v):
            doms.append(v)

def _parse_csv_or_xlsx(file_bytes: bytes, filename: str) -> Tuple[List[str], List[str], List[List[str]]]:
    ips, doms = [], []
    preview_rows: List[List[str]] = []

    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8", "ignore")
        sample = "\n".join(text.splitlines()[:50]) or ""
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        for idx, row in enumerate(reader):
            if len(row) > 1 and all(cell.isdigit() for cell in row):
                row = [",".join(row)]
            row = ["" if c is None else str(c).strip() for c in row]
            if idx < 10:
                preview_rows.append(row)
            _auto_detect_cells(row, ips, doms)
    else:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(values_only=True)):
            row = ["" if c is None else str(c).strip() for c in r]
            if idx < 10:
                preview_rows.append(row)
            _auto_detect_cells(row, ips, doms)

    ips = sorted(set(ips))
    doms = sorted(set(doms))
    return ips, doms, preview_rows

@app.post("/api/upload")
async def upload_file(action: str = Form(..., pattern="^(preview|process)$"), file: UploadFile = File(...)):
    content = await file.read()
    ips, doms, preview = _parse_csv_or_xlsx(content, file.filename)
    if action == "preview":
        return {"filename": file.filename, "detected": {"ips": len(ips), "domains": len(doms)}, "preview": preview[:10]}
    payload = {"ips": ips, "domains": doms}
    return await lookup_batch(payload)

# ===================== IA (com timeout defensivo) ===================== #
@app.get("/api/ai-status")
async def ai_status():
    return {"enabled": bool(getattr(settings, "ENABLE_AI", False))}

@app.post("/api/insights")
async def insights(payload: dict = Body(...)):
    try:
        if not getattr(settings, "ENABLE_AI", False):
            return JSONResponse({"detail": "IA desativada pelo administrador."}, status_code=403)
        if not getattr(settings, "OPENAI_API_KEY", None):
            return JSONResponse({"detail": "OPENAI_API_KEY ausente nas configurações."}, status_code=500)

        ctx = payload.get("context") or {}
        q = (ctx.get("query") or "").strip()
        t = (ctx.get("type") or "").strip().lower()
        if not q or t not in {"ip", "domain", "hash"}:
            return JSONResponse({"detail": "Contexto inválido: informe 'type' (ip|domain|hash) e 'query'."}, status_code=400)

        client = OpenAI(api_key=settings.OPENAI_API_KEY)

        base_instructions = """
Você é um Analista Sênior de Threat Intelligence. Escreva em Português, com objetividade e formato Markdown.
Estruture SEMPRE em:
1) Visão Geral
2) Evidências Relevantes (cite os pontos extraídos das fontes)
3) Avaliação de Risco (Baixo/Médio/Alto) — justificando com dados
4) Recomendações Ação-Orientadas (curto prazo e médio prazo)
5) Lacunas de Dados (o que faltou para aumentar a confiança)
Não invente dados: use somente o que vier no 'input' e a experiência comum de resposta a incidentes.
"""
        hash_focus = """
Contexto: O artefato a ser analisado é um HASH (MD5/SHA256) potencialmente malicioso.
Objetivo: Forneça um parecer que ajude resposta a incidentes e hunting.
Inclua, quando possível a partir das fontes:
- VirusTotal (files): malicious/suspicious/harmless/undetected, type_description, nomes de detecção comuns.
- Hybrid Analysis: verdict, threat_score, vx_family, tipo e quaisquer IOCs comportamentais.
- Indicadores acionáveis (IOCs): nomes de arquivo, caminhos, URLs, domínios, endereços, mutex/registry, relações.
- Risco: postura conservadora para múltiplas detecções ou verdicts “malicious/suspicious”.
- Contenção/Erradicação: quarentena por hash, bloqueio no EDR/AV, varredura em endpoints.
- Monitoramento/Hunting: YARA de alto nível (esqueleto) e buscas SIEM/Sigma de alto nível.
- Boas práticas: não executar amostras; manuseio seguro.
Se algo não estiver disponível, sinalize em "Lacunas de Dados".
"""
        instructions = base_instructions + (hash_focus if t == "hash" else "")

        input_payload = {
            "tipo": t,
            "consulta": q,
            "fontes": ctx.get("sources") or {}
        }

        # Tempo máximo do endpoint (default 60s se não houver em config.py)
        overall_deadline = int(getattr(settings, "AI_MAX_SECONDS", 60))

        async def _call_openai():
            # chamada síncrona do client em thread p/ não travar o loop
            def sync_call():
                resp = client.responses.create(
                    model=getattr(settings, "OPENAI_MODEL", "gpt-4o-mini"),
                    instructions=instructions,
                    input=json.dumps(input_payload, indent=2, ensure_ascii=False)
                )
                return getattr(resp, "output_text", None) or "Sem conteúdo gerado."
            return await asyncio.to_thread(sync_call)

        try:
            text = await asyncio.wait_for(_call_openai(), timeout=overall_deadline)
        except asyncio.TimeoutError:
            return JSONResponse(
                {"detail": "AI demorou para responder", "timeout_seconds": overall_deadline},
                status_code=504,
            )

        return JSONResponse({"insights": text}, status_code=200)

    except Exception as e:
        return JSONResponse({"detail": f"Falha ao gerar insights: {e}"}, status_code=500)

# ===================== utils ===================== #
async def _gather_safe(*aws):
    out = []
    for c in aws:
        try:
            out.append(await c)
        except httpx.HTTPStatusError as e:
            out.append({"error": f"HTTP {e.response.status_code}", "detail": e.response.text})
        except httpx.RequestError as e:
            out.append({"error": "request_error", "detail": str(e)})
        except Exception as e:
            out.append({"error": "unknown_error", "detail": str(e)})
    return out

